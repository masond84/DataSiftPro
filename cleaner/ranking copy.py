# import modules
import pandas as pd
from openpyxl import *


#### Get Keyword Weights ####
def get_keyword_weights(filepath):
    """ 
    Given a file (excel or csv) extract the columns keywords: weights
    and file them into a dictionary with keywords as keys and weights as the value
    {'keyword': weight}
    """
    try:
        print(f"Reading File: {filepath}")
        # Read the file using pandas, depending on the file type
        df = pd.read_csv(filepath) if filepath.endswith('.csv') else pd.read_excel(filepath)
        
        # Check if the 'Keyword' and 'Weight' columns are present in the file
        if 'Keyword' in df.columns and 'Weight' in df.columns:
            # Create an empyty dictionary to store the keyword weights
            keyword_weights = {}
            
            # Loop through each row of the dataframe
            for _, row in df.iterrows():
                # Get the keyword (lowercase) and its weight from the row
                keyword = row['Keyword'].lower()
                weight = row['Weight']

                # Add the keyword and its weight to the dictionary
                keyword_weights[str(keyword)] = int(weight)
            
            # Return the dictionary of keyword weights
            return keyword_weights
        else:
            # Raise a value error if either 'Keyword' or 'Weight' columns are missing
            raise ValueError("File does not contain both 'Keyword' and 'Weight' columns.")
    except Exception as e:
    # Print an error message if there's an issue reading the file
        print("Error Reading File: ", e)
        # Return an empty dictionary if there's an error
        return {}

def count_keywords(text, keyword_dict):
    """
    Given a string of text and a dictionary of keywords, loop through the text
    and count how many times each keyword appears. Returns the updated keyword
    dictionary with the counts.
    """
    for keyword in keyword_dict:
        # count how many times the keyword appears in the text
        count = text.lower().count(keyword.lower())
        # Add the count to the value of the keyword key
        keyword_dict[keyword] += count   
    return keyword_dict

#### Score Lists ####
'''name_scores = []
n_keywords = []
desc_scores = []
desc_keywords = []
industry_scores = []
industry_keywords = []
specialty_scores = []
specialty_keywords = []
employee_scores = []'''

def rank_company(row, keywords, thresholds, scores):
    # Overall score for each individual company
    score = 0
    # assign the dictionary of keywords/weights 
    weights = keywords
    
    #### CHECK COMPANY NAME FOR KEYWORDS ####
    name_score = 0 # individual namescore values
    # Keep track of how many times each keyword appears when searched for
    all_keys = {}
    for word in keywords: 
        all_keys[word] = 0
    # Company Name - Dataframe
    company_name = str(row["Company Name"]).lower().replace(',', '').replace('.', '')
    #####################################
    counts = count_keywords(company_name, all_keys)
    #####################################
    # loop through the keywords and their weights
    for keyword, weight in weights.items():
        # Check if the keyword appears in the company name
        if keyword in company_name:
            # Add to the individual name score value
            name_score += weight
            # Increment the overall company score by weighted keyword value
            score += weight
        else: 
            if " " in keyword and keyword in company_name:
                name_score += weight
                all_keys[keyword] += 1
    # Delete Keywords That Were Not Found #
    for key in list(all_keys.keys()):
        if all_keys[key] == 0:
            del all_keys[key]
    # Name Score Value Text
    name_text = (f"{name_score}")
    name_scores.append(name_text)
    # Name Score Keyword Text
    name_text = (f"{counts}")
    n_keywords.append(name_text)
    
    #### CHECK COMPANY DESCRIPTION FOR KEYWORDS ####
    desc_score = 0 # individual description score values
    # keep track of how many times each keyword appears when searched for
    all_keys = {}
    for word in keywords: 
        all_keys[word] = 0
    description = str(row['Description']).lower().replace(',', '').replace('.', '').replace('(', '').replace(')', '')    
    ############################################
    counts = count_keywords(description, all_keys)
    ###########################################
    for keyword, weight in weights.items():
        if keyword in description:
            desc_score += weight
            score += weight
    # Delete Keywords That Were Not Found #
    for key in list(all_keys.keys()):
        if all_keys[key] == 0:
            del all_keys[key]
    # Description Score Value Text
    desc_text = (f"{desc_score}")
    desc_scores.append(desc_text)
    # Description Score Keyword Text
    desc_text = (f"{counts}")
    desc_keywords.append(desc_text)
    
    #### CHECK COMPANY INDUSTRY FOR KEYWORDS ####
    industry_score = 0
    # keep track of how many times each keyword appears when searched for
    all_keys = {}
    for word in keywords: 
        all_keys[word] = 0
    if isinstance(row['Industries'], str):
        industry = row['Industries'].lower().replace(',', '')
        ######################################
        counts = count_keywords(industry, all_keys)
        ######################################
        for keyword, weight in weights.items():
            if keyword in industry:
                industry_score += weight
                score += weight
    # Delete Keywords That Were Not Found #
    for key in list(all_keys.keys()):
        if all_keys[key] == 0:
            del all_keys[key]
    # Industry Score Value Text
    industry_text = (f"{industry_score}")
    industry_scores.append(industry_text)
    # Industry Score Keyword Text
    industry_text = (f"{counts}")
    industry_keywords.append(industry_text)

    #### CHECK COMPANY SPECIALTY FOR KEYWORDS ####
    specialty_score = 0
    # keep track of how many times each keyword appears when searched for
    all_keys = {}
    for word in keywords: 
        all_keys[word] = 0
    if isinstance(row['Specialties'], str):
        specialties = row['Specialties'].lower().replace(',', '')
        ######################################
        counts = count_keywords(specialties, all_keys)
        ######################################
        for keyword, weight in weights.items():
            if keyword in specialties:
                specialty_score += weight
                score += weight
    # Delete Keywords That Were Not Found #
    for key in list(all_keys.keys()):
        if all_keys[key] == 0:
            del all_keys[key]
    # Specialty Score Value Text
    specialty_text = (f"{specialty_score}")
    specialty_scores.append(specialty_text)
    # Specialty Score Keyword Text
    specialty_text = (f"{counts}")
    specialty_keywords.append(specialty_text)

    ''' Revise and Edit how we rank according to employee count'''
    #### CHECK EMPLOYEE COUNT TO ADD TO SCORE ####
    # dynamic employee count function
    employee_score = 0
    #print('running employee score')
    for i in range(len(thresholds)):
        if row['Employee Count'] >= thresholds[i]:
            employee_score = scores[i]
            break
        #print(f'Adding {employee_score} to employee score')
        score += employee_score
        employee_text = (f"{employee_score}")
        employee_scores.append(employee_text)
    
    # return the final score value of each company
    return score

def rank_companies(filepath, keyword_file, thresholds, scores, company_count=1000000):
    global name_scores, n_keywords, desc_scores, desc_keywords, industry_scores, industry_keywords, specialty_scores, specialty_keywords, employee_scores
    #### Score Lists ####
    name_scores = []
    n_keywords = []
    desc_scores = []
    desc_keywords = []
    industry_scores = []
    industry_keywords = []
    specialty_scores = []
    specialty_keywords = []
    employee_scores = []
    # Get the sourcedata file
    FNAME = filepath
    # Get the specified keywords
    keywords = get_keyword_weights(keyword_file)

    # Load the workbook
    workbook = load_workbook(FNAME)
    ws = workbook.active
    # dataframe
    df = pd.read_excel(FNAME)
    
    #### CREATE A NEW SHEET FOR RANKED COMPANIES ####
    if "Company Ranking" in workbook.sheetnames:
        ws = workbook["Company Ranking"]
    else:
        ws = workbook.create_sheet("Company Ranking")

    ### Create New Columns Below ###
    # apply the ranking function to each row of the dataframe
    df['Score'] = df.apply(lambda x: rank_company(x, keywords, thresholds, scores), axis=1)
    df['Name Score'] = name_scores
    df['Name Keywords'] = n_keywords
    df['Description Score'] = desc_scores
    df['Description Keywords'] = desc_keywords
    df['Industry Score'] = industry_scores
    df['Industry Keywords'] = industry_keywords
    df['Specialty Score'] = specialty_scores
    df['Specialty Keywords'] = specialty_keywords
    
    # Convert employee_scores to a DataFrame
    employee_scores_df = pd.DataFrame(employee_scores, columns=['Employee Score'])

    # Find the indices in df that are missing in employee_scores_df
    missing_indices = set(df.index) - set(employee_scores_df.index)

    # Filter the main DataFrame to get the rows with missing scores
    missing_scores = df.loc[list(missing_indices)]

    print("Companies missing scores:", missing_scores['Company Name'].tolist())

    # apply the individual scores 
    df.to_excel(FNAME, index=False)
    # sort the companies by score in descending order
    df_sorted = df.sort_values('Score', ascending=False)
    # add the 
    # select the top 50-100 companies
    company_rank = company_count # specify the amount of companies you want to display
    condensed_list = df_sorted.head(company_rank)


    # Check the column names and assign the correct one
    if "Latest Estimated Revenue" in df.columns:
        revenue_header = "Latest Estimated Revenue ($)"
    elif "Latest Estimated Revenue ($)" in df.columns:
        revenue_header = "Latest Estimated Revenue ($)"
    else:
        raise KeyError("Column name does not exist in source file: Latest Estimated Revenue")
    
    # UPDATED ROW ('Latest Estimated Revenue ($)')
    header = ["Score", "Name Score", "Name Keywords", "Company Name",
            "Description Score", "Description Keywords", "Description",
            "Industry Score", "Industry Keywords", "Industries",
            "Specialty Score", "Specialty Keywords", "Specialties",
            "Employee Score", "Employee Count", "Website", 
            "City", "State", "Postal Code", "Country",
            "Phone Number", "LinkedIn Account", "Informal Name", "Founding Year",
            "Employee Range", "3 Months Growth Rate %", "6 Months Growth Rate %", "9 Months Growth Rate %", "12 Months Growth Rate %",
            "Growth Intent", "Ownership", "Total Raised", "Latest Raised", "Date of Most recent Investment", "Investors",
            "Parent Company", "Executive Title", "Executive First Name", "Executive Last Name", "Executive Email",
            "Executive LinkedIn", "Last Financial Year", "Verified Revenue", f"{revenue_header}", "Financial Growth %", "Financial Growth Period (yr)", 
            "Sources Count", "CRM Id", "My Tags", "Firm Tags"
            ]
    ws.append(header)
        
    # print the top 50-100 companies
    for i, row in condensed_list.iterrows():
        # print the index of that row in the dataframe, the name of the company, and the company score
        #print(f"{row['Company Name']}: {row['Website']}: Score: {row['Score']}")
        # Add Company Headers
        company_data = [row["Score"], row["Name Score"], row["Name Keywords"], row["Company Name"],
                        row["Description Score"], row["Description Keywords"], row["Description"],
                        row["Industry Score"], row["Industry Keywords"], row["Industries"],
                        row["Specialty Score"], row["Specialty Keywords"], row["Specialties"],
                        row["Employee Score"], row["Employee Count"], row["Website"],
                        row["City"], row["State"], row["Postal Code"], row["Country"],
                        row["Phone Number"], row["LinkedIn Account"], row["Informal Name"], row["Founding Year"],
                        row["Employee Range"], row["3 Months Growth Rate %"], row["6 Months Growth Rate %"], row["9 Months Growth Rate %"], row["12 Months Growth Rate %"],
                        row["Growth Intent"], row["Ownership"], row["Total Raised"], row["Latest Raised"], row["Date of Most recent Investment"], row["Investors"],
                        row["Parent Company"], row["Executive Title"], row["Executive First Name"], row["Executive Last Name"], row["Executive Email"],
                        row["Executive LinkedIn"], row["Last Financial Year"], row["Verified Revenue"], row[revenue_header], row["Financial Growth %"], row["Financial Growth Period (yr)"],
                        row["Sources Count"], row["CRM Id"], row["My Tags"], row["Firm Tags"]
                        ] 
        ws.append(company_data)

    #### CREATE A NEW SHEET FOR KEYWORDS AND WEIGHTS ####
    if "Keyword Weights" in workbook.sheetnames:
        pass
    else:
        workbook.create_sheet("Keyword Weights")
    ws = workbook["Keyword Weights"]
    # Assign a header
    header = ["Keywords", "Weights"]
    ws.append(header)
    
    # Create a dataframe from the dictionary
    df = pd.DataFrame.from_dict(keywords, orient="index", columns=["Weights"])

    for i, row in df.iterrows():
        ws.append([i, row['Weights']])
    # Write the dataframe to an excel file
    #df.to_excel(FNAME, index=False)

    # Save the workbook
    ranked_data = workbook.save(FNAME)
    
    return (ranked_data)